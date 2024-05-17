import openpyxl 
import datetime

print("打开：MT料况表.xlsx")
SitWorkbook = openpyxl.load_workbook('MT料况表.xlsx',data_only=True)

SitSheet = (SitWorkbook["电子料"],
            SitWorkbook["结构料"],
            SitWorkbook["包材"])


# 创建一个新的Excel文件
NewWorkbook = openpyxl.Workbook()
# 选择一个工作表
NewSheet = (NewWorkbook.create_sheet("电子料",0),
           NewWorkbook.create_sheet("结构料",1),
           NewWorkbook.create_sheet("包材",2))


SpaceSize = 0
for i in range(1,SitSheet[0].max_row+1):
    if(SitSheet[0].cell(i, 1).value!=None):
        SpaceSize=i
        break
offset = 0
for i in range(1,SitSheet[0].max_column+1):
    if(SitSheet[0].cell(SpaceSize, i).value=="替代料"):
        offset = i-1
        break

print ("查找数据量")
LineEnd = 0
for i in range(3):
    sheetMax_row = SitSheet[i].max_row-(SpaceSize-1)
    sheetMax_column=SitSheet[i].max_column-((offset+1)*2)

    SitSheet[i].delete_cols(1,offset+1)
    SitSheet[i].delete_cols(4,offset+1)
    SitSheet[i].delete_rows(1,SpaceSize-1)
    for j in range(0,SitSheet[i].max_column-1):
        if(isinstance(SitSheet[i].cell(1, SitSheet[i].max_column-j).value, datetime.datetime)):
            LineEnd=int(SitSheet[i].max_column-j)
            break
    NewTraversals=int((SitSheet[i].max_row-(2))/3)
    NewMaxrow = SitSheet[i].max_row
    for j in range(0,NewTraversals):
        print("查缺料：",NewTraversals,j)
        if((SitSheet[i].cell(NewMaxrow-j*3, LineEnd).value==None)or(int(SitSheet[i].cell(NewMaxrow-j*3, LineEnd).value)>=0)):
            SitSheet[i].delete_rows(NewMaxrow-j*3-2,3)
            sheetMax_row-=3
    for j in range(1,sheetMax_row+1):#行遍历↓
        for k in range(1,sheetMax_column+1):#列遍历→
            NewSheet[i].cell(j,k).value=SitSheet[i].cell(j,k).value
    DatePosition=sheetMax_column+1
    databufold=0
    InitialPosition=0
    for k in range(1,sheetMax_column+1):#列遍历→
        if(k>=DatePosition):
            NewSheet[i].cell(1,k).number_format = 'm/d'
        elif(NewSheet[i].cell(1,k).value=="共用料"):
            DatePosition=k
            
            for j in range(3,sheetMax_row+1):#行遍历↓
                if(NewSheet[i].cell(j,k).value!=None):
                    databuf1=NewSheet[i].cell(j,k).value
                if(databufold!=databuf1):
                    databufold = databuf1
                    if(InitialPosition!=0):
                        setbuf=(openpyxl.utils.get_column_letter(k)+str(InitialPosition)+":"+openpyxl.utils.get_column_letter(k)+str(j-1))
                        NewSheet[i].merge_cells(setbuf)
                    InitialPosition=j
            else:
                try:
                    setbuf=(openpyxl.utils.get_column_letter(k)+str(InitialPosition)+":"+openpyxl.utils.get_column_letter(k)+str(j))
                    NewSheet[i].merge_cells(setbuf)
                except:
                    print("i=",i, "j=",j, "k=",k, "setbuf=",setbuf,"InitialPosition=",InitialPosition,"sheetMax_row=",sheetMax_row,"databufold=",databufold)
            
NewWorkbook.save(".\MT缺料表.xlsx")
NewWorkbook.close()
SitWorkbook.close()
