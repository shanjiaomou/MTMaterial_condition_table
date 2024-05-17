import os
import openpyxl

# 指定文件夹路径
folder_path = './bom'

# 获取文件夹中的所有文件和文件夹
file_list = os.listdir(folder_path)

#设置单元格格式
cell_format = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
#设置单元格边
thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'), 
                                             right=openpyxl.styles.borders.Side(style='thin'), 
                                             top=openpyxl.styles.borders.Side(style='thin'), 
                                             bottom=openpyxl.styles.borders.Side(style='thin'))

# 创建一个新的Excel文件
workbook = openpyxl.Workbook()
# 选择一个工作表
sheet = workbook.active
#设置行高
sheet.row_dimensions[1].height=36.8
sheet.row_dimensions[2].height=77.3

arr = [
    ["替代料","小米料号","迈腾代码","物料描述","使用比例"],
    [4,15,18,9,4]
]
#增加一列用于相同替代料排序
offset = len(file_list)+1
arrange = len(file_list)+1
for i in range(5):
    sheet.cell(1, arrange).value = arr[0][i]
    sheet.cell(1, arrange).border = thin_border
    sheet.merge_cells(openpyxl.utils.get_column_letter(arrange)+"1:"+openpyxl.utils.get_column_letter(arrange)+"2")
    sheet.column_dimensions[openpyxl.utils.get_column_letter(arrange)].width = arr[1][i]
    sheet.cell(1, arrange).alignment = cell_format
    arrange+=1

Entry_Line = 3  #总表行数
column = 1
Index_number = 0    #索引号
Index_number_total = 0 #索引号排序
Index_End = 0       #最后索引行
Temporary_Start = 0 #临时区开始
# 遍历列表并输出 文件层
for file_name in file_list:
    print("开始处理",file_name)
    #打开excel表
    ReadWorkbook = openpyxl.load_workbook(folder_path+'/'+file_name)
    #通过工作表名称选择工作表
    worksheet = ReadWorkbook['2 PCBA']
    #读取料号
    cell_value = worksheet['N3'].value
    #获取型号
    Model_Last=file_name.find(".xlsx")
    if Model_Last>0:
        Model_Name = file_name[20:Model_Last]
    else:
        Model_Name = '格式错误'
    #设置列宽
    sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 4
    sheet.column_dimensions[openpyxl.utils.get_column_letter(arrange)].width = 4
    #填入数据
    sheet.cell(1, column).value = cell_value            #料号
    sheet.cell(1, column).alignment = cell_format       #居中 自动换行
    sheet.cell(1, column).border = thin_border          #边框
    sheet.cell(2, column).value = Model_Name            #型号
    sheet.cell(2, column).alignment = cell_format       #居中 自动换行
    sheet.cell(2, column).border = thin_border          #边框
    sheet.cell(1, arrange).value = cell_value           #料号
    sheet.cell(1, arrange).alignment = cell_format      #居中 自动换行
    sheet.cell(1, arrange).border = thin_border         #边框
    sheet.cell(2, arrange).value = Model_Name           #型号
    sheet.cell(2, arrange).alignment = cell_format      #居中 自动换行
    sheet.cell(2, arrange).border = thin_border         #边框
    
    max_row = worksheet.max_row-3 #bom行数
    workcount = 4   #从第四行开始遍历
    databuf1o1 = 0      #旧料号1
    databuf1o2 = 0      #旧料号2
    #bom料号层
    while(workcount<max_row):
        databuf1=int(worksheet.cell(workcount, 1).value)/10 #bom获取料编号

        if(databuf1o2!=databuf1):#处理新料前 要把临时区先处理
            databuf1o2 = databuf1
            if(Temporary_Start!=0):#有临时区吗
                if(Index_End!=0):#索引值
                    Index_End+=1
                    for i in range(Entry_Line-Temporary_Start):
                        sheet.insert_rows((Index_End), 1)
                        sheet.cell(Index_End, column).value = sheet.cell(Entry_Line, column).value#写入bom料编号
                        sheet.cell(Index_End, offset+1).value = sheet.cell(Entry_Line, offset+1).value #写入料号
                        sheet.cell(Index_End, offset+3).value = sheet.cell(Entry_Line, offset+3).value #写描述
                        sheet.cell(Index_End, offset+4).value = sheet.cell(Entry_Line, offset+4).value  #料权等于1
                        sheet.cell(Index_End, arrange).value = sheet.cell(Entry_Line, arrange).value #写入用量
                        sheet.cell(Index_End, offset).value = Index_number
                        sheet.delete_rows(Entry_Line)
                else:
                    Index_number_total+=1
                    for i in range(Entry_Line-Temporary_Start):
                        sheet.cell(Temporary_Start+i, offset).value = Index_number_total
            Index_number = 0
            Index_End = 0
            Temporary_Start = 0
        sheetcount = 3  #总表第三行开始索引
        #总表索引
        while(sheetcount<Entry_Line):
            #判断料号一致
            if(worksheet.cell(workcount, 14).value==sheet.cell(sheetcount, offset+1).value):
                sheet.cell(sheetcount, column).value = databuf1 #写入bom料编号
                #sheet.cell(sheetcount, offset+4).value = int(sheet.cell(sheetcount, offset+4).value)+int(worksheet.cell(workcount, 19).value) #写入总用量
                sheet.cell(sheetcount, arrange).value = worksheet.cell(workcount, 19).value #写入用量
                Index_number = sheet.cell(sheetcount, offset).value
                break
            if(Index_number!=0):
                if(Index_number == sheet.cell(sheetcount, offset).value):
                    if(Index_End<sheetcount):
                        Index_End = sheetcount
            sheetcount+=1
        #新物料
        else:
            if(databuf1o1!=databuf1):#新料
                databuf1o1 = databuf1
                Temporary_Start = sheetcount
            sheet.cell(sheetcount, column).value = databuf1 #写入bom料编号
            sheet.cell(sheetcount, offset+1).value = worksheet.cell(workcount, 14).value #写入料号
            sheet.cell(sheetcount, offset+3).value = worksheet.cell(workcount, 17).value #写描述
            sheet.cell(sheetcount, offset+4).value = 0 #worksheet.cell(workcount, 19).value #写入总用量
            sheet.cell(sheetcount, arrange).value = worksheet.cell(workcount, 19).value #写入用量
            Entry_Line+=1
        workcount+=1    #bom料跳到第二行
    #关闭Excel文件
    ReadWorkbook.close()
    column += 1
    arrange+= 1
if(Temporary_Start!=0):#有临时区吗
    if(Index_End!=0):#索引值
        Index_End+=1
        for i in range(Entry_Line-Temporary_Start):
            sheet.insert_rows((Index_End), 1)
            sheet.cell(Index_End, column).value = sheet.cell(Entry_Line, column).value#写入bom料编号
            sheet.cell(Index_End, offset+1).value = sheet.cell(Entry_Line, offset+1).value #写入料号
            sheet.cell(Index_End, offset+3).value = sheet.cell(Entry_Line, offset+3).value #写描述
            sheet.cell(Index_End, offset+4).value = sheet.cell(Entry_Line, offset+4).value  #料权等于1
            sheet.cell(Index_End, arrange).value = sheet.cell(Entry_Line, arrange).value #写入用量
            sheet.cell(Index_End, offset).value = Index_number
            sheet.delete_rows(Entry_Line)
    else:
        Index_number_total+=1
        for i in range(Entry_Line-Temporary_Start):
            sheet.cell(Temporary_Start+i, offset).value = Index_number_total
try:
    # 保存文件
    workbook.save(".\00.BOM整合清单.xlsx")
    
    print("整合BOM完成")
except:
    print("保存BOM失败,请勿操作文档")
print("开始设置使用比例")
databuf1o1 = 0
Index_Start = 0 #索引行开始
for sheetcount in range(3,(sheet.max_row+1)):
    databuf1 = sheet.cell(sheetcount, offset).value
    if(databuf1o1!=databuf1):
        sheet.cell(sheetcount, offset+4).value = 1  #设置权重
        Index_Start = sheetcount
        databuf1o1 = databuf1
    else:
        for i in range(offset+5,offset*2+5):
            if(sheet.cell(sheetcount,i).value is None):
                pass
            else:
                for j in range(Index_Start,sheetcount):
                    if(sheet.cell(j,i).value is None):
                        pass
                    else:
                        break
                else:
                    #删重复
                    for k in range(offset+5,offset*2+5):
                        for j in range(Index_Start,sheetcount):
                            if(sheet.cell(j,k).value is None):
                                pass
                            else:
                                sheet.cell(sheetcount, k).value = None
                                break
                    sheet.cell(sheetcount, offset+4).value = 1 #设置权重
                    break
    print(".",end="")
print('.')

try:
    # 保存文件
    workbook.save(".\BOM使用比例清单.xlsx")
    workbook.close()
    print("保存使用比例完成")
except:
    print("保存使用比例失败,请勿操作文档")



