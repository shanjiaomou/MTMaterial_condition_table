import openpyxl 
import random

#设置单元格格式
cell_format = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
#设置单元格边
thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'), 
                                             right=openpyxl.styles.borders.Side(style='thin'), 
                                             top=openpyxl.styles.borders.Side(style='thin'), 
                                             bottom=openpyxl.styles.borders.Side(style='thin'))

# 打开Excel文件
BomWorkbook = openpyxl.load_workbook('BOM使用比例清单.xlsx',data_only=True)
try:
    NeedWorkbook = openpyxl.load_workbook('小米生产主计划.xlsx',data_only=True)

except:
    print("请准备文件：小米生产主计划.xlsx")
    exit()

NewWorkbook = openpyxl.Workbook()   #新表格

SpaceSize=2

# 选择一个工作表
BomSheet = (BomWorkbook["电子料"],
            BomWorkbook["结构料"],
            BomWorkbook["包材"])
NewSheet = (NewWorkbook.create_sheet("电子料",0),
            NewWorkbook.create_sheet("结构料",1),
            NewWorkbook.create_sheet("包材",2))
NeedSheet = NeedWorkbook["小米生产计划"]
#NeedSheet.calculate()
if(NeedSheet.cell(4, 3).value!="整机料号")\
or(NeedSheet.cell(4, 4).value!="半成品料号")\
or(NeedSheet.cell(4, 5).value!="PCBA"):
    print("计划表格式不对！")
    exit()

TypesStr = ["电子料","结构料","包材"]

PlannedHeight=NeedSheet.max_row-3

offset = 0
for i in range(1,BomSheet[0].max_column+1):
    if(BomSheet[0].cell(1, i).value=="替代料"):
        offset = i-1
        break

for i in range(3):
    print("开始处理",TypesStr[i])
    print("  开始拷贝计划单")
    NewSheet[i].cell(1,offset*2+5+SpaceSize).value = NeedSheet.cell(4, 6).value     #迈腾成品料号 4 6
    NewSheet[i].cell(1,offset*2+5+SpaceSize+1).value = NeedSheet.cell(4, 5-i).value #整机/半成品/PCBA
    NewSheet[i].cell(1,offset*2+5+SpaceSize+2).value = NeedSheet.cell(4, 2).value   #机型

    for j in range(8,12):
        NewSheet[i].cell(1,offset*2+SpaceSize+j).value = NeedSheet.cell(4, j).value
    
    for j in range(offset*2+5+SpaceSize,offset*2+12+SpaceSize):
        NewSheet[i].cell(1, j).border = thin_border #边框
        NewSheet[i].column_dimensions[openpyxl.utils.get_column_letter(j)].width = 10 #列宽
        NewSheet[i].cell(1, j).alignment = cell_format #居中
        NewSheet[i].merge_cells(openpyxl.utils.get_column_letter(j)+"1:"+openpyxl.utils.get_column_letter(j)+"2")


    for j in range(12,NeedSheet.max_column+1):
        NewSheet[i].column_dimensions[openpyxl.utils.get_column_letter(offset*2+SpaceSize+j)].width = 10 #列宽
        NewSheet[i].cell(1, offset*2+SpaceSize+j).value = NeedSheet.cell(4, j).value
        NewSheet[i].cell(1, offset*2+SpaceSize+j).border = thin_border #边框
        NewSheet[i].cell(1, offset*2+SpaceSize+j).alignment = cell_format #居中
        NewSheet[i].cell(1, offset*2+SpaceSize+j).number_format = 'm月d日'
        NewSheet[i].cell(2, offset*2+SpaceSize+j).value = NeedSheet.cell(5, j).value
        NewSheet[i].cell(2, offset*2+SpaceSize+j).border = thin_border #边框
        NewSheet[i].cell(2, offset*2+SpaceSize+j).alignment = cell_format #居中
        NewSheet[i].cell(2, offset*2+SpaceSize+j).number_format = '(aaa)'
    NewSheet[i].row_dimensions[1].height = 12.8 #设置行高
    NewSheet[i].row_dimensions[2].height = 12.8 #设置行高

    for k in range(6,NeedSheet.max_row+1):
        NewSheet[i].cell(k-3,offset*2+5+SpaceSize).value = NeedSheet.cell(k, 6).value     #迈腾成品料号 4 6
        NewSheet[i].cell(k-3,offset*2+5+SpaceSize+1).value = NeedSheet.cell(k, 5-i).value #整机/半成品/PCBA
        NewSheet[i].cell(k-3,offset*2+5+SpaceSize+2).value = NeedSheet.cell(k, 2).value   #机型
        for j in range(3):
            NewSheet[i].cell(k-3, offset*2+5+SpaceSize+j).border = thin_border #边框
            NewSheet[i].cell(k-3, offset*2+5+SpaceSize+j).alignment = cell_format #居中
        for j in range(8,NeedSheet.max_column+1):
            NewSheet[i].cell(k-3,offset*2+SpaceSize+j).value = NeedSheet.cell(k, j).value
            NewSheet[i].cell(k-3, offset*2+SpaceSize+j).border = thin_border #边框
            NewSheet[i].cell(k-3, offset*2+SpaceSize+j).alignment = cell_format #居中
        NewSheet[i].row_dimensions[k-3].height = 12.8 #设置行高
    print("  开始拷贝bom表")
    #复制bom表
    for j in range(1,BomSheet[i].max_column+1):
        if((j<=(offset+1))or((j>=(offset+5))and(j<=(offset*2+6)))):
            NewSheet[i].column_dimensions[openpyxl.utils.get_column_letter(j)].width = 4
        elif(j==(offset+2)):
            NewSheet[i].column_dimensions[openpyxl.utils.get_column_letter(j)].width = 15
        elif(j==(offset+3)):
            NewSheet[i].column_dimensions[openpyxl.utils.get_column_letter(j)].width = 18
        elif(j==(offset+4)):
            NewSheet[i].column_dimensions[openpyxl.utils.get_column_letter(j)].width = 9
        for k in range(1,BomSheet[i].max_row+1):
            NewSheet[i].cell(k+PlannedHeight,j).value = BomSheet[i].cell(k,j).value
            NewSheet[i].cell(k+PlannedHeight,j).border = thin_border #边框
            NewSheet[i].cell(k+PlannedHeight,j).alignment = cell_format #居中
        if(((j>=(offset+1))and(j<=(offset+5)))or(j==(offset*2+6))):#合并单元格
            setbuf=(openpyxl.utils.get_column_letter(j)+str(PlannedHeight+1)+":"+openpyxl.utils.get_column_letter(j)+str(PlannedHeight+2))
            NewSheet[i].merge_cells(setbuf)

    NewSheet[i].cell(1+PlannedHeight,offset*2+10+SpaceSize).value = "共用料"
    NewSheet[i].cell(1+PlannedHeight,offset*2+10+SpaceSize).border = thin_border #边框
    NewSheet[i].cell(1+PlannedHeight,offset*2+10+SpaceSize).alignment = cell_format #居中
    setbuf=(openpyxl.utils.get_column_letter(offset*2+10+SpaceSize)+str(PlannedHeight+1)+":"+openpyxl.utils.get_column_letter(offset*2+10+SpaceSize)+str(PlannedHeight+2))
    NewSheet[i].merge_cells(setbuf)
    NewSheet[i].cell(1+PlannedHeight,offset*2+11+SpaceSize).value = "区分"
    NewSheet[i].cell(1+PlannedHeight,offset*2+11+SpaceSize).border = thin_border #边框
    NewSheet[i].cell(1+PlannedHeight,offset*2+11+SpaceSize).alignment = cell_format #居中
    setbuf=(openpyxl.utils.get_column_letter(offset*2+11+SpaceSize)+str(PlannedHeight+1)+":"+openpyxl.utils.get_column_letter(offset*2+11+SpaceSize)+str(PlannedHeight+2))
    NewSheet[i].merge_cells(setbuf)
    for j in range(offset*2+12+SpaceSize,NewSheet[i].max_column+1):
        NewSheet[i].cell(1+PlannedHeight,j).value = NewSheet[i].cell(1,j).value
        NewSheet[i].cell(1+PlannedHeight,j).border = thin_border #边框
        NewSheet[i].cell(1+PlannedHeight,j).alignment = cell_format #居中
        NewSheet[i].cell(1+PlannedHeight,j).number_format = 'm/d'
        setbuf=(openpyxl.utils.get_column_letter(j)+str(PlannedHeight+1)+":"+openpyxl.utils.get_column_letter(j)+str(PlannedHeight+2))
        NewSheet[i].merge_cells(setbuf)
    OldMaterials=0
    StartMaterials=0
    
    print("  开始赋需求初值")
    #索引采购订单对应的型号
    for k in range(3+PlannedHeight,NewSheet[i].max_row+1):#遍历物料↓
        ''' 共用料
        if((NewSheet[0].cell(k,offset+1).value!=None)and(OldMaterials!=NewSheet[0].cell(k,offset+1).value)):
            OldMaterials=NewSheet[0].cell(k,offset+1).value
            NewSheet[i].cell(k,offset*2+SpaceSize+10).border = thin_border #边框
            NewSheet[i].cell(k,offset*2+SpaceSize+10).alignment = cell_format #居中
            if(StartMaterials!=0):#合并单元格
                NewSheet[i].cell(StartMaterials,offset*2+SpaceSize+10).value=OldMaterials-1
                setbuf=(openpyxl.utils.get_column_letter(offset*2+SpaceSize+10)+str(StartMaterials)+":"+openpyxl.utils.get_column_letter(offset*2+SpaceSize+10)+str(k-1))
                NewSheet[i].merge_cells(setbuf)
            StartMaterials=k
        '''
        NewSheet[i].cell(k,offset*2+SpaceSize+11).value = "需求"
        """
        NewSheet[i].cell(k,offset*2+SpaceSize+11).border = thin_border #边框
        NewSheet[i].cell(k,offset*2+SpaceSize+11).alignment = cell_format #居中
        """
        for l in range(offset*2+SpaceSize+12,NewSheet[i].max_column+1):#遍历日期→
            NewSheet[i].cell(k,l).value = 0
            """
            NewSheet[i].cell(k,l).border = thin_border #边框
            NewSheet[i].cell(k,l).alignment = cell_format #居中 
            """
    print("  开始计算需求")
    for k in range(3,PlannedHeight+1):#遍历计划型号↓
        for l in range(offset+6,offset*2+5):#遍历BOM型号→
            if(NewSheet[i].cell(k,offset*2+6+SpaceSize).value==NewSheet[i].cell(1+PlannedHeight,l).value):#匹配型号
                #匹配了，按日期算值
                for m in range(offset*2+SpaceSize+12,NewSheet[i].max_column+1):#遍历日期→
                    if(NewSheet[i].cell(k,m).value!=None):#判断当天有生产
                        for o in range(3+PlannedHeight,NewSheet[i].max_row+1):#遍历物料↓
                            if(NewSheet[i].cell(o,l).value!=None):#判断bom料用量
                                OriginalValue=int(NewSheet[i].cell(o,m).value)#获取当天之前bom用量
                                UseRatio=int(NewSheet[i].cell(o,offset+5).value)
                                CalculatedValue=int(NewSheet[i].cell(k,m).value)*int(NewSheet[i].cell(o,l).value)
                                NewSheet[i].cell(o,m).value=OriginalValue+UseRatio*CalculatedValue
                break
    print("  开始插入ETA&GAP")
    rowend=NewSheet[i].max_row


    for j in range(0,rowend-PlannedHeight-2):
        NewSheet[i].insert_rows((PlannedHeight+4+j*3), 2)#插入2行

       
        ''' 
        for k in range(1,offset*2+SpaceSize+12): #合并单元格，设置
            NewSheet[i].cell(k,(rowend+1-j)-4).border = thin_border #边框
            NewSheet[i].cell(k,(rowend+1-j)-4).alignment = cell_format #居中
            setbuf=(openpyxl.utils.get_column_letter(k)+str((rowend+1-j)-4)+":"+openpyxl.utils.get_column_letter(k)+str((rowend+1-j)-1))
            NewSheet[i].merge_cells(setbuf)
            '''
        print("插入处理第:",(rowend-PlannedHeight-2),":",j,"行")

    for k in range(3+PlannedHeight,NewSheet[i].max_row+1):#遍历物料↓
        print("GAP处理第:",(NewSheet[i].max_row+1),":",k,"行")
        if(NewSheet[i].cell(k,offset*2+11+SpaceSize).value == "需求"):
            NewSheet[i].cell(k+1,offset*2+11+SpaceSize).value = "ETA"
            NewSheet[i].cell(k+2,offset*2+11+SpaceSize).value = "GAP"
            RequirementLocation=openpyxl.utils.get_column_letter(offset*2+12+SpaceSize)+str(k)
            ETALocation=openpyxl.utils.get_column_letter(offset*2+12+SpaceSize)+str(k+1)
            InventoryLocation=openpyxl.utils.get_column_letter(offset*2+6)+str(k)
            NewSheet[i].cell(k+2,offset*2+12+SpaceSize).value = "="+InventoryLocation+"+"+ETALocation+"-"+RequirementLocation
            for j in range(offset*2+12+SpaceSize+1,NewSheet[i].max_column+1):
                RequirementLocation=openpyxl.utils.get_column_letter(j)+str(k)
                ETALocation=openpyxl.utils.get_column_letter(j)+str(k+1)
                InventoryLocation=openpyxl.utils.get_column_letter(j-1)+str(k+2)
                NewSheet[i].cell(k+2,j).value = "="+InventoryLocation+"+"+ETALocation+"-"+RequirementLocation
            if((NewSheet[i].cell(k,offset+1).value!=None)and(OldMaterials!=NewSheet[i].cell(k,offset+1).value)):
                OldMaterials=NewSheet[i].cell(k,offset+1).value
                ''' 共用料
                NewSheet[i].cell(k,offset*2+SpaceSize+10).border = thin_border #边框
                NewSheet[i].cell(k,offset*2+SpaceSize+10).alignment = cell_format #居中
                '''
                if(StartMaterials!=0):#合并单元格
                    NewSheet[i].cell(StartMaterials,offset*2+SpaceSize+10).value=OldMaterials-1
                    setbuf=(openpyxl.utils.get_column_letter(offset*2+SpaceSize+10)+str(StartMaterials)+":"+openpyxl.utils.get_column_letter(offset*2+SpaceSize+10)+str(k-1))
                    NewSheet[i].merge_cells(setbuf)
                StartMaterials=k

try:
    # 保存文件
    NewWorkbook.save(".\MT料况表.xlsx")
    print("保存："+".\MT料况表.xlsx")
except:
    SaveName=".\MT料况表"+str(random.randint(0,65535))+".xlsx"
    NewWorkbook.save(SaveName)
    print("保存："+SaveName)


NewWorkbook.close()
