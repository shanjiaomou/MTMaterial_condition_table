"""
本代码由[Tkinter布局助手]生成
官网:https://www.pytk.net/tkinter-helper
QQ交流群:788392508
"""
import os
import openpyxl
import datetime
from tkinter import filedialog
from tkinter import messagebox
from tkinter import *
from tkinter.ttk import *
from typing import Dict

#设置数组数据
def array_set(array, row, col, value):
    # 获取二维数组的行数和列数
    rows = len(array)
    cols = len(array[0])
    # 如果插入位置超出数组范围，则扩展数组并使用空值补全
    if col >= cols:
        for i in range(cols, col ):
            for i in range(rows):
                array[i].insert(col, None)
        # 在指定位置插入新列
        for i in range(rows):
            array[i].insert(col, None)
        # 获取二维数组的行数和列数
        cols = len(array[0])
    values=[None for _ in range(cols)]
    # 如果插入位置超出数组范围，则扩展数组并使用空值补全
    if row >= rows:
        for i in range(rows, row ):
            array.insert(i, values)
        # 在指定位置插入新行
        array.insert(row, values)
    # 在指定位置插入值
    array[row][col] = value
    return array
#获取数组数据
def array_get(array, row, col):
    # 获取二维数组的行数和列数
    rows = len(array)   #行
    cols = len(array[0])#列
    # 如果插入位置超出数组范围，则扩展数组并使用空值补全
    if row >= rows or col >= cols:
        return False
    return array[row][col]
#插入行
def insert_row(array, row_index):
    # 获取二维数组的行数和列数
    rows = len(array)
    cols = len(array[0])
    values=[None for _ in range(cols)]
    # 如果插入位置超出数组范围，则扩展数组并使用空值补全
    if row_index > rows:
        for i in range(rows, row_index ):
            array.insert(i, values)
    # 在指定位置插入新行
    array.insert(row_index, values)
    return array
#插入列
def insert_col(array, col_index):
    # 获取二维数组的行数和列数
    rows = len(array)
    cols = len(array[0])

    # 如果插入位置超出数组范围，则扩展数组并使用空值补全
    if col_index > cols:
        for i in range(cols, col_index ):
            for i in range(rows):
                array[i].insert(col_index, None)
    # 在指定位置插入新列
    for i in range(rows):
        array[i].insert(col_index, None)
    return array   
#插入剪切行
def CutInsert_row(array,origin,finish):
    # 将第origin行剪切插入第finish行前面
    row_to_insert = array.pop(origin)  # 移除第origin行并保存
    array.insert(finish, row_to_insert)  # 在第finish行前面插入移除的行
    return array
#插入剪切列
def CutInsert_col(array,origin,finish):
    # 将第origin列剪切插入第finish列前面
    for row in array:
        col_to_insert = row.pop(origin)  # 移除第origin列并保存
        row.insert(finish, col_to_insert)  # 在第finish列前面插入移除的列
    return array
#获取行数
def array_row(array):
    return len(array)
#获取列数
def array_col(array):
    return len(array[0])
#错误弹窗
def show_error(srtbuf):
    messagebox.showerror("错误", srtbuf)
def print_label(self,srtbuf):
    self.tk_label_lm0opp7u["text"]=srtbuf
def print_log(self,srtbuf):
    self.tk_text_lm0p7c95.configure(state="normal")
    self.tk_text_lm0p7c95.insert(END,srtbuf+"\r\n")
    self.tk_text_lm0p7c95.configure(state="disabled")
    self.tk_text_lm0p7c95.see(END)
def Clear_log(self):
    self.tk_text_lm0p7c95.configure(state="normal")
    self.tk_text_lm0p7c95.delete("1.0", END)
    self.tk_text_lm0p7c95.configure(state="disabled")
def set_Prog(self,value,maximum):
    self.tk_progressbar_lm0p6l8w["maximum"]=maximum
    self.tk_progressbar_lm0p6l8w["value"]=value
#App#######################################################################
#用下面函数处理数据

def Sheet_Handle(SheetBuf,array):
    cell_value = SheetBuf['N3'].value                                       #获取主件料号（固定位置）
    ArrayCol = array_col(array)                                             #获取数组列数
    ArrayItem = ArrayCol-2                                                  #序号在表格中的列
    ArrayDos = ArrayCol-1                                                   #用量在表格中的列
    array = array_set(array,0,ArrayDos, cell_value)                         #写入主件料号
    array = array_set(array,0,ArrayItem,cell_value)                         #写入主件料号    
    BomMaxRow = SheetBuf.max_row                                            #获取物料工作表行数
    OldItem = 0                                                             #初始化旧序号
    SMNum = 0                                                               #初始化替代料编号缓存
    SMAddr = 0                                                              #初始化替代料末尾地址
    BufNum = 0                                                              #初始化缓存数量
    ArrayRow = array_row(array)                                             #获取数组行数
    if(ArrayRow>2):                                                         #判断有数据了
        SMMax = array_get(array,ArrayRow-1,0)                               #赋值替代料编号计数
    else:                                                                   #否则
        SMMax = 0                                                           #初始化替代料编号计数
    for sheet_i in range(4,BomMaxRow+1):                                    #遍历处理物料
        if((SheetBuf.cell(sheet_i, 13).value=="P")                          
           and(SheetBuf.cell(sheet_i, 22).value=="Y")):                     #判断物料有效，采购层次、认证情况
            PartItem = int(SheetBuf.cell(sheet_i, 1).value)/10              #获取序号（固定位置），用于区分共用料
            PartNO=SheetBuf.cell(sheet_i, 14).value                         #获取料号（固定位置）
            Description=SheetBuf.cell(sheet_i, 17).value                    #获取描述（固定位置）
            Dosage = SheetBuf.cell(sheet_i, 19).value                       #获取用量（固定位置）
            ArrayRow = array_row(array)                                     #获取数组行数
            if(OldItem!=PartItem):
                if(BufNum!=0):
                    SMMax+=1
                    for BufNum_i in range(BufNum):
                        array = array_set(array,ArrayRow-1-BufNum_i,0,SMMax) #更新共用关系值
                SMNum = 0
                SMAddr = 0
                BufNum = 0
                OldItem = PartItem
            
            for array_i in range(2,ArrayRow):           #遍历总bom
                if(array_get(array,array_i,1)==PartNO): #料号相同
                    array = array_set(array,array_i,ArrayItem,PartItem)     #写入序号
                    array = array_set(array,array_i,ArrayDos, Dosage)       #写入用量
                    setbuf1 = array_get(array,array_i,3)                     #获取适用型号
                    setbuf2 = array_get(array,1, ArrayItem).replace(" 序号", "")#获取当前型号
                    Model_Name = setbuf1+","+setbuf2
                    array = array_set(array,array_i,3,Model_Name)           #设置适用型号
                    
                    if((SMNum!=0)and(SMAddr!=0)):#有共用关系
                        SMBuf = array_get(array,array_i,0)
                        if(SMNum != SMBuf):#共用关系不同
                            if(SMAddr>array_i):#老的大
                                for i in range(1,SMAddr-1):
                                    if(SMNum != array_get(array,SMAddr-i,0)):
                                        SMBufStart = SMAddr-i+1
                                        break
                                    else:
                                        SMBufStart = 2
                                SMBufEnd = SMAddr
                                SMNum = SMBuf
                                for array_j in range(array_i,ArrayRow):
                                    if(SMNum != array_get(array,array_j,0)):
                                        SMAddr = array_j
                                        break
                                else:
                                    SMAddr = ArrayRow
                            else:#新的大
                                for i in range(array_i-1):
                                    if(SMBuf != array_get(array,array_i-i,0)):
                                        SMBufStart = array_i-i+1
                                        break
                                    else:
                                        SMBufStart = 2
                                for i in range(array_i,ArrayRow):
                                    if(SMBuf != array_get(array,i,0)):
                                        SMBufEnd = i
                                        break
                                    else:
                                        SMBufEnd = ArrayRow
                            SMBufLen = SMBufEnd-SMBufStart
                            for SMBuf_i in range(SMBufLen):
                                array = array_set(array,SMBufEnd-1,0,SMNum) #更新共用关系值
                                array = CutInsert_row(array,SMBufEnd-1,SMAddr) #移动数据
                            SMAddr+= SMBufLen #更新共用地址  
                    else: #无共用关系
                        SMNum = array_get(array,array_i,0)#获取共用关系值
                        for array_j in range(array_i,ArrayRow):
                            if(SMNum != array_get(array,array_j,0)):
                                SMAddr = array_j
                                break
                        else:
                            SMAddr = ArrayRow
                        if(BufNum!=0):
                            for BufNum_i in range(BufNum):
                                array = array_set(array,ArrayRow-1,0,SMNum) #更新共用关系值
                                array = CutInsert_row(array,ArrayRow-1,SMAddr) #移动数据
                            SMAddr+= BufNum #更新共用地址
                            BufNum = 0      #情况缓存标志
                    break
            else:                                       #料号不同
                if((SMNum!=0)and(SMAddr!=0)):#有共用关系
                    AddrBuf = SMAddr
                    array = insert_row(array, AddrBuf) #插入行
                    array = array_set(array,AddrBuf,0,SMNum) #更新共用关系值
                    SMAddr+=1
                else: #无共用关系
                    BufNum+=1
                    AddrBuf = ArrayRow
                array = array_set(array,AddrBuf,1,          PartNO)         #写入料号
                array = array_set(array,AddrBuf,2,          Description)    #写入描述
                array = array_set(array,AddrBuf,ArrayItem,  PartItem)       #写入序号
                array = array_set(array,AddrBuf,ArrayDos,   Dosage)         #写入用量
                
                Model_Name = array_get(array,1, ArrayItem).replace(" 序号", "")#获取当前型号
                array = array_set(array,AddrBuf,3,          Model_Name)     #设置适用型号
    else:
        if(BufNum!=0):
            ArrayRow = array_row(array)
            SMMax+=1
            for BufNum_i in range(BufNum):
                array = array_set(array,ArrayRow-1-BufNum_i,0,SMMax) #更新共用关系值
    return array

def bom_format(array):
    ArrayCol = array_col(array)
    ModelCol = int((ArrayCol-4)/2)
    for i in range(ModelCol):
        array = CutInsert_col(array,(ArrayCol-2-i),0)
    return array

def bom_UseRatio(array):
    ArrayCol = array_col(array)
    ModelCol = int((ArrayCol-4)/2)                          #替代料位置
    array=insert_col(array,ModelCol+4)
    array=array_set(array, 0, ModelCol+4, "使用比例")       #插入使用比例列
    array=insert_col(array,ModelCol+2)
    array=array_set(array, 0, ModelCol+2, "迈腾代码")       #插入迈腾代码列
    RatioLie = ModelCol+5                                   #使用比例位置
    MateStart = 0                                           #初始化索引行开始
    ArrayRow = array_row(array)
    OldItem = 0                                             #初始化旧序号
    ArrayCol = array_col(array)                             #获取总列数
    ItemStart = 0
    ItemEnd = 0
    for row_i in range(2,ArrayRow):                         #行遍历
        array = array_set(array,row_i,RatioLie,0)
    for row_i in range(2,ArrayRow+1):
        PartItem = array_get(array,row_i,ModelCol)          #获取序号
        if(OldItem!=PartItem):                              #序号变更
            ItemStart = ItemEnd
            ItemEnd = row_i
            OldItem = PartItem
            if((ItemStart!=0)and(ItemEnd!=0)):
                for col_i in range(ModelCol):
                    ModeDict = dict()
                    for row_j in range(ItemStart,ItemEnd):
                        SMNum = array_get(array,row_j,col_i) #获取编号
                        if(SMNum!=None):
                            if(ModeDict.get(SMNum, None)!=None): #字典里有？
                                if(array_get(array,row_j,RatioLie)==1): #使用比例=1？
                                    array = array_set(array,row_j,RatioLie+1+col_i,0)#清除用量
                            else:
                                ModeDict[SMNum]=row_j
                                if(array_get(array,row_j,RatioLie)==0):
                                    array = array_set(array,row_j,RatioLie,1)
                                    for col_j in range(col_i):
                                        if((array_get(array,row_j,RatioLie+1+col_j)!=None)and(array_get(array,row_j,RatioLie+1+col_j)!=0)):
                                            array = array_set(array,row_j,RatioLie+1+col_j,0)#清除用量
    for row_i in range(2,ArrayRow):                         #行遍历
        if(array_get(array,row_i,RatioLie)==0):
            for col_i in range(ModelCol):
                if(array_get(array,row_i,RatioLie+1+col_i)!=None):
                    array = array_set(array,row_i,RatioLie+1+col_i,0) #设置用量
            array = array_set(array,row_i,RatioLie,1) #设置使用比例
    for col_i in range(ModelCol+1):
        array = CutInsert_col(array,ArrayCol-1,ModelCol)
    return array 

def main_app(self):
    #创建数组，增加PCBA、结构料、包材 三个表的标题 
    TotalArray = [[["序号","小米料号","物料描述","项目号"],[None,None,None,None]],
                  [["序号","小米料号","物料描述","项目号"],[None,None,None,None]],
                  [["序号","小米料号","物料描述","项目号"],[None,None,None,None]]]
    Clear_log(self)
    # 获取文件夹中的所有文件和文件夹
    folder_path = self.tk_input_lm0omywa.get()
    try:
        file_list = os.listdir(folder_path)
    except:
        show_error("请选择正确BOM目录")
        return
    file_len = len(file_list)
    #一个一个遍历所有文件
    for file_count in range(file_len):
        set_Prog(self,file_count+1,file_len)#更新进度条
        file_name = file_list[file_count]
        print_label(self,"导入BOM:"+str(file_count+1)+"/"+str(file_len))
        print_log(self,"导入BOM:"+file_name)
        Model_Last=file_name.find(".xlsx")
        if(("BOM清单_L95"!=file_name[0:9])or(0>Model_Last)):
            print_log(self," └文件名错误")
            win.update()#更新界面
            continue
        win.update()#更新界面
        # 产品型号
        Model_Name = file_name[20:Model_Last]
        # 加载Excel文档
        BomWorkBook=openpyxl.load_workbook(folder_path+"/"+file_name,data_only=True)
        # 获取所有工作表的名称
        BomSheetNames = BomWorkBook.sheetnames
        # 遍历输出每个工作表的名称
        for BomSheetName in BomSheetNames:
            BomSheet = BomWorkBook[BomSheetName]
            if(0<BomSheetName.find("PACKAGING")):
                #获取列号（成品列）
                ArrayCol      = array_col(TotalArray[2])
                TotalArray[2] = array_set(TotalArray[2],1,ArrayCol,  Model_Name+" 序号")
                TotalArray[2] = array_set(TotalArray[2],1,ArrayCol+1,Model_Name+" 用量")
                TotalArray[2] = Sheet_Handle(BomSheet,TotalArray[2])
            elif(0<BomSheetName.find("PCBA")):
                #获取列号（成品列）
                ArrayCol      = array_col(TotalArray[0])
                TotalArray[0] = array_set(TotalArray[0],1,ArrayCol,  Model_Name+" 序号")
                TotalArray[0] = array_set(TotalArray[0],1,ArrayCol+1,Model_Name+" 用量")                
                TotalArray[0] = Sheet_Handle(BomSheet,TotalArray[0])
            elif(0<BomSheetName.find("FA")):
                #获取列号（成品列）
                ArrayCol      = array_col(TotalArray[1])
                TotalArray[1] = array_set(TotalArray[1],1,ArrayCol,  Model_Name+" 序号")
                TotalArray[1] = array_set(TotalArray[1],1,ArrayCol+1,Model_Name+" 用量")
                TotalArray[1] = Sheet_Handle(BomSheet,TotalArray[1])
            else:
                print_log(self," └错误工作页："+BomSheetName)
                win.update()#更新界面
                show_error(file_name+"\r\n工作页错误:"+BomSheetName+"\r\n请排查")
        #释放处理完的工作簿
        BomWorkBook.close()
    for i in range(3):
        ArrayRow = array_row(TotalArray[i])
        OldItem = 0
        SMNum = 0
        for row_i in range(2,ArrayRow):
            PartItem = array_get(TotalArray[i],row_i,0)
            if(OldItem!=PartItem):
                SMNum += 1
                OldItem = PartItem
            TotalArray[i] = array_set(TotalArray[i],row_i,0,SMNum)
    print_label(self,"调整格式")
    print_log(self,"调整格式")
    set_Prog(self,0,3)#更新进度条
    win.update()#更新界面
    #调整数组位置
    for i in range(3):
        set_Prog(self,i+1,3)#更新进度条
        win.update()#更新界面
        TotalArray[i]=bom_format(TotalArray[i])

    TotalArrayLen = [array_row(TotalArray[0]),array_row(TotalArray[1]),array_row(TotalArray[2])]
    ProgLen = TotalArrayLen[0]+TotalArrayLen[1]+TotalArrayLen[2]
    set_Prog(self,0,ProgLen)#更新进度条
    print_label(self,"存储BOM整合清单")
    print_log(self,"存储BOM整合清单")
    win.update()#更新界面
    # 创建一个新的Excel文件
    OutWorkBook = openpyxl.Workbook()
    # 选择一个工作表
    OutSheet = (OutWorkBook.create_sheet("电子料",0),
               OutWorkBook.create_sheet("结构料",1),
               OutWorkBook.create_sheet("包材",2))
    for i in range(3):
        for j in range(TotalArrayLen[i]):
            if i == 0:
                ProgBar = j
            elif i == 1:
                ProgBar = TotalArrayLen[0]+j
            else :
                ProgBar = TotalArrayLen[0]+TotalArrayLen[0]+j
            set_Prog(self,ProgBar,ProgLen)#更新进度条
            win.update()#更新界面
            OutSheet[i].append(TotalArray[i][j])
    try:
        # 获取当前日期
        current_date = str(datetime.date.today().strftime("%Y%m%d"))
        SaveName=".\BOM整合清单"+current_date+".xlsx"
        OutWorkBook.save(SaveName)
    except:
        current_datetime = str(datetime.datetime.now().strftime("%Y%m%d%H%M%S"))
        SaveName=".\BOM整合清单"+current_datetime+".xlsx"
        OutWorkBook.save(SaveName)
    OutWorkBook.close
    print_label(self,"保存："+SaveName)
    print_log(self,"保存："+SaveName)
    win.update()#更新界面
    
    # 创建一个新的Excel文件
    OutWorkBook = openpyxl.Workbook()
    # 选择一个工作表
    OutSheet = (OutWorkBook.create_sheet("电子料",0),
               OutWorkBook.create_sheet("结构料",1),
               OutWorkBook.create_sheet("包材",2))    
    for i in range(3):
        print_label(self,"处理使用比例："+str(i+1)+"/3")
        print_log(self,"处理使用比例："+str(i+1)+"/3")
        win.update()#更新界面
        TotalArray[i] = bom_UseRatio(TotalArray[i])
        for j in range(TotalArrayLen[i]):
            if i == 0:
                ProgBar = j
            elif i == 1:
                ProgBar = TotalArrayLen[0]+j
            else :
                ProgBar = TotalArrayLen[0]+TotalArrayLen[0]+j
            set_Prog(self,ProgBar,ProgLen)#更新进度条
            win.update()#更新界面
            OutSheet[i].append(TotalArray[i][j])
    try:
        current_date = str(datetime.date.today().strftime("%Y%m%d"))
        SaveName=".\BOM使用比例清单"+current_date+".xlsx"
        OutWorkBook.save(SaveName)
    except:
        current_datetime = str(datetime.datetime.now().strftime("%Y%m%d%H%M%S"))
        SaveName=".\BOM使用比例清单"+current_datetime+".xlsx"
        OutWorkBook.save(SaveName)
    OutWorkBook.close
    print_label(self,"保存："+SaveName)
    print_log(self,"保存："+SaveName)
    win.update()#更新界面
    
#UI#######################################################################
class WinGUI(Tk):
    def __init__(self):
        super().__init__()
        self.__win()
        self.tk_label_lm0omf5i = self.__tk_label_lm0omf5i(self)
        self.tk_input_lm0omywa = self.__tk_input_lm0omywa(self)
        self.tk_button_lm0on4cw = self.__tk_button_lm0on4cw(self)
        self.tk_button_lm0onmq9 = self.__tk_button_lm0onmq9(self)
        self.tk_label_lm0opp7u = self.__tk_label_lm0opp7u(self)
        self.tk_progressbar_lm0p6l8w = self.__tk_progressbar_lm0p6l8w(self)
        self.tk_text_lm0p7c95 = self.__tk_text_lm0p7c95(self)
    def __win(self):
        self.title("BOM整合比例工具V2.3")
        # 设置窗口大小、居中
        width = 420
        height = 240
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        geometry = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(geometry)
        self.resizable(width=False, height=False)
        # 自动隐藏滚动条
    def scrollbar_autohide(self,bar,widget):
        self.__scrollbar_hide(bar,widget)
        widget.bind("<Enter>", lambda e: self.__scrollbar_show(bar,widget))
        bar.bind("<Enter>", lambda e: self.__scrollbar_show(bar,widget))
        widget.bind("<Leave>", lambda e: self.__scrollbar_hide(bar,widget))
        bar.bind("<Leave>", lambda e: self.__scrollbar_hide(bar,widget))
    
    def __scrollbar_show(self,bar,widget):
        bar.lift(widget)
    def __scrollbar_hide(self,bar,widget):
        bar.lower(widget)
    
    def vbar(self,ele, x, y, w, h, parent):
        sw = 15 # Scrollbar 宽度
        x = x + w - sw
        vbar = Scrollbar(parent)
        ele.configure(yscrollcommand=vbar.set)
        vbar.config(command=ele.yview)
        vbar.place(x=x, y=y, width=sw, height=h)
        self.scrollbar_autohide(vbar,ele)
    def __tk_label_lm0omf5i(self,parent):
        label = Label(parent,text="BOM目录：",anchor="center", )
        label.place(x=20, y=20, width=70, height=30)
        return label
    def __tk_input_lm0omywa(self,parent):
        ipt = Entry(parent, )
        ipt.place(x=90, y=20, width=310, height=30)
        ipt.insert(0, (os.path.dirname(os.path.abspath(__file__)))+'\BOM') # 获取.py文件所在的目录路径
        return ipt
    def __tk_button_lm0on4cw(self,parent):
        btn = Button(parent, text="选择目录", takefocus=False,)
        btn.place(x=260, y=60, width=60, height=30)
        return btn
    def __tk_button_lm0onmq9(self,parent):
        btn = Button(parent, text="开始处理", takefocus=False,)
        btn.place(x=340, y=60, width=60, height=30)
        return btn
    def __tk_label_lm0opp7u(self,parent):
        label = Label(parent,text="请选择文件夹并开始",anchor="w" )
        label.place(x=20, y=60, width=220, height=30)
        return label
    def __tk_progressbar_lm0p6l8w(self,parent):
        progressbar = Progressbar(parent, orient=HORIZONTAL,)
        progressbar.place(x=20, y=100, width=380, height=10)
        return progressbar
    def __tk_text_lm0p7c95(self,parent):
        text = Text(parent,state="disabled")
        text.place(x=20, y=120, width=380, height=100)
        self.vbar(text, 20, 120, 380, 100,parent)
        return text
class Win(WinGUI):
    def __init__(self):
        super().__init__()
        self.__event_bind()
    def OpenFolderEvent(self):#(self,evt):
        # 打开文件夹选择对话框，设置初始目录为.py文件所在的位置
        folder_path = filedialog.askdirectory(initialdir=self.tk_input_lm0omywa.get())  
        if(folder_path!=""):
            self.tk_input_lm0omywa.delete(0, END)  # 清空输入框内容
            self.tk_input_lm0omywa.insert(END, folder_path)  # 将选择的目录路径填入输入框
    def StartProcessEvent(self):#(self,evt):
        self.tk_button_lm0onmq9.config(state=DISABLED)
        main_app(self)
        self.tk_button_lm0onmq9.config(state=NORMAL)
    def __event_bind(self):
        #self.tk_button_lm0on4cw.bind('<Button>',self.OpenFolderEvent)
        #self.tk_button_lm0onmq9.bind('<Button>',self.StartProcessEvent)
        self.tk_button_lm0on4cw["command"]=self.OpenFolderEvent
        self.tk_button_lm0onmq9["command"]=self.StartProcessEvent
        pass
if __name__ == "__main__":
    win = Win()
    win.mainloop()