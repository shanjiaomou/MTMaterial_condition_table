import openpyxl 
import datetime

print("打开：MT料况表.xlsx")
NewWorkbook = openpyxl.load_workbook('MT料况表.xlsx',data_only=True)

NewSheet = (NewWorkbook["电子料"],
            NewWorkbook["结构料"],
            NewWorkbook["包材"])

SpaceSize = 0
for i in range(1,NewSheet[0].max_row+1):
    if(NewSheet[0].cell(i, 1).value!=None):
        SpaceSize=i
        break
offset = 0
for i in range(1,NewSheet[0].max_column+1):
    if(NewSheet[0].cell(SpaceSize, i).value=="替代料"):
        offset = i-1
        break

print ("查找数据量")
LineEnd = 0
for i in range(3):
    for j in range(0,NewSheet[i].max_column-1):
        #if(isinstance(NewSheet[i].cell(SpaceSize, NewSheet[i].max_column-j).value, openpyxl.cell.cell.Cell)):
        if(isinstance(NewSheet[i].cell(SpaceSize, NewSheet[i].max_column-j).value, datetime.datetime)):
            LineEnd=int(NewSheet[i].max_column-j)
            break
    NewTraversals=int((NewSheet[i].max_row-(SpaceSize+1))/3)
    NewMaxrow = NewSheet[i].max_row
    for j in range(0,NewTraversals):
        print("查缺料：",NewTraversals,j)
        if((NewSheet[i].cell(NewMaxrow-j*3, LineEnd).value==None)or(int(NewSheet[i].cell(NewMaxrow-j*3, LineEnd).value)>=0)):
            NewSheet[i].delete_rows(NewMaxrow-j*3-2,3)
#        if(isinstance(NewSheet[i].cell(SpaceSize, NewSheet[i].max_column-j).value, int)):
#            print(openpyxl.utils.get_column_letter(NewSheet[i].max_column-j))
#            break
    NewSheet[i].delete_cols(1,offset+1)
    NewSheet[i].delete_cols(4,offset+1)
    NewSheet[i].delete_rows(1,SpaceSize-1)

NewWorkbook.save(".\MT缺料表.xlsx")

NewWorkbook.close()
