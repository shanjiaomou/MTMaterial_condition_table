"""
本代码由[Tkinter布局助手]生成
官网:https://www.pytk.net/tkinter-helper
QQ交流群:788392508
"""
import os
import openpyxl
import locale
import datetime
from tkinter import filedialog
from tkinter import messagebox
from tkinter import *
from tkinter.ttk import *
from typing import Dict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 设置地区为中国
locale.setlocale(locale.LC_ALL, 'zh_CN.UTF-8')

def CxTransform(row, col):
    # 将数字列号转换为字母列号
    col_letter = get_column_letter(col)
    # 构建字符串坐标
    cell_coordinate = f"{col_letter}{row}"
    return cell_coordinate

def excel_to_date(excel_timestamp):
    # 将数字日期值转换为日期对象
    excel_start_date = datetime.datetime(1899, 12, 30)  # Excel将数字日期从1899年12月30日开始计数
    delta = datetime.timedelta(days=excel_timestamp)  # 创建一个时间间隔，表示从起始日期开始的天数
    date_value = excel_start_date + delta  # 将起始日期和时间间隔相加得到最终的日期值
    # 返回结果，日期和星期分别使用strftime()函数格式化输出
    return date_value.strftime('%m月%d日'),date_value.strftime('%m/%d'), date_value.strftime('%A')
    
#设置数组数据
def array_set(array, row, col, value):
    # 获取二维数组的行数和列数
    rows = len(array)
    cols = len(array[0])
    # 如果插入位置超出数组范围，则扩展数组并使用空值补全
    if col >= cols:
        for i in range(cols, col ):
            for i in range(rows):
                array[i].insert(col, None)
        # 在指定位置插入新列
        for i in range(rows):
            array[i].insert(col, None)
        # 获取二维数组的行数和列数
        cols = len(array[0])
    values=[None for _ in range(cols)]
    # 如果插入位置超出数组范围，则扩展数组并使用空值补全
    if row >= rows:
        for i in range(rows, row ):
            array.insert(i, values)
        # 在指定位置插入新行
        array.insert(row, values)
    # 在指定位置插入值
    array[row][col] = value
    return array
#获取数组数据
def array_get(array, row, col):
    # 获取二维数组的行数和列数
    rows = len(array)   #行
    cols = len(array[0])#列
    # 如果插入位置超出数组范围，则扩展数组并使用空值补全
    if row >= rows or col >= cols:
        return False
    return array[row][col]
#插入行
def insert_row(array, row_index):
    # 获取二维数组的行数和列数
    rows = len(array)
    cols = len(array[0])
    values=[None for _ in range(cols)]
    # 如果插入位置超出数组范围，则扩展数组并使用空值补全
    if row_index > rows:
        for i in range(rows, row_index ):
            array.insert(i, values)
    # 在指定位置插入新行
    array.insert(row_index, values)
    return array
#插入列
def insert_col(array, col_index):
    # 获取二维数组的行数和列数
    rows = len(array)
    cols = len(array[0])

    # 如果插入位置超出数组范围，则扩展数组并使用空值补全
    if col_index > cols:
        for i in range(cols, col_index ):
            for i in range(rows):
                array[i].insert(col_index, None)
    # 在指定位置插入新列
    for i in range(rows):
        array[i].insert(col_index, None)
    return array   
#插入剪切行
def CutInsert_row(array,origin,finish):
    # 将第origin行剪切插入第finish行前面
    row_to_insert = array.pop(origin)  # 移除第origin行并保存
    array.insert(finish, row_to_insert)  # 在第finish行前面插入移除的行
    return array
#插入剪切列
def CutInsert_col(array,origin,finish):
    # 将第origin列剪切插入第finish列前面
    for row in array:
        col_to_insert = row.pop(origin)  # 移除第origin列并保存
        row.insert(finish, col_to_insert)  # 在第finish列前面插入移除的列
    return array
#获取行数
def array_row(array):
    return len(array)
#获取列数
def array_col(array):
    return len(array[0])
#UI界面
def show_error(srtbuf):
    messagebox.showerror("错误", srtbuf)
def print_label(self,srtbuf):
    self.tk_label_lm3lw72f["text"]=srtbuf
def print_log(self,srtbuf):
    self.tk_text_lm3m3ylm.configure(state="normal")
    self.tk_text_lm3m3ylm.insert(END,srtbuf+"\r\n")
    self.tk_text_lm3m3ylm.configure(state="disabled")
    self.tk_text_lm3m3ylm.see(END)
def Clear_log(self):
    self.tk_text_lm3m3ylm.configure(state="normal")
    self.tk_text_lm3m3ylm.delete("1.0", END)
    self.tk_text_lm3m3ylm.configure(state="disabled")
def set_Prog(self,value,maximum):
    self.tk_progressbar_lm3lze9x["maximum"]=maximum
    self.tk_progressbar_lm3lze9x["value"]=value
#App#######################################################################

def main_app(self):
    Clear_log(self)
    TotalArray = [[],
                  [],
                  []]
    Bom_path = self.tk_input_lm3ln8cp.get()
    Plan_path = self.tk_input_lm3lqahh.get()
    if(os.path.isfile(Bom_path)==False):
        show_error("请正确选择BOM清单!")
        return
    if(os.path.isfile(Plan_path)==False):
        show_error("请正确选择计划!")
        return
    
    Plan_name = os.path.basename(Plan_path)
    print_label(self,"导入计划表:\""+Plan_name+"\"中...")
    print_log(self,"导入计划表:"+Plan_name)
    set_Prog(self,1,3)
    win.update()#更新界面
    try:
        PlanWorkbook = openpyxl.load_workbook(Plan_path,data_only=True)
        PlanSheet = PlanWorkbook["小米生产计划"]
    except:
        show_error("打开\""+Plan_name+"\"失败!"+"\r\n请检查文件是否正常!")        
        return 
    Bom_name = os.path.basename(Bom_path)
    print_label(self,"导入BOM清单:\""+Bom_name+"\"中...")
    print_log(self,"导入BOM清单:"+Bom_name)
    set_Prog(self,2,3)
    win.update()#更新界面
    try:
        BomWorkbook = openpyxl.load_workbook(Bom_name,data_only=True)
        BomSheet = (BomWorkbook["电子料"],
                    BomWorkbook["结构料"],
                    BomWorkbook["包材"])
    except:
        show_error("打开\""+Bom_name+"\"失败!"+"\r\n请检查文件是否正常!")        
        return 

    TypeList = [5, 4, 3]#计划数量
    PlanMaxRow = [2, 2, 2]#计划数量
    if((PlanSheet.cell(4,TypeList[0]).value!="PCBA")
       or(PlanSheet.cell(4,TypeList[1]).value!="半成品料号")
       or(PlanSheet.cell(4,TypeList[2]).value!="整机料号")):
        show_error("计划表格式错误,请检查!")
        return
    for i in range(3):
        print_label(self,"预处理BOM清单:"+str(i+1)+"/3")
        print_log(self,"预处理BOM清单:"+str(i+1)+"/3")
        MaxRow=BomSheet[i].max_row
        row_i = 0
        for row in BomSheet[i].iter_rows(values_only=True):
            set_Prog(self,row_i,MaxRow)
            win.update()#更新界面
            row_i +=1 
            TotalArray[i].append(list(row))
        ArrayCol = array_col(TotalArray[i])
        UsePosition = int((ArrayCol-6)/2)
        BomWorkbook.close()
        if(array_get(TotalArray[i],0,UsePosition)!="使用比例"):
            show_error("找不到使用比例啦!")
            return
        TotalArray[i]=insert_col(TotalArray[i],ArrayCol)
        TotalArray[i]=array_set(TotalArray[i], 0, ArrayCol, "物料状态") #插入
        TotalArray[i]=insert_col(TotalArray[i],ArrayCol)
        TotalArray[i]=array_set(TotalArray[i], 0, ArrayCol, "缺料时间") #插入
        TotalArray[i]=insert_col(TotalArray[i],ArrayCol)
        TotalArray[i]=array_set(TotalArray[i], 0, ArrayCol, "可用库存") #插入
        TotalArray[i]=insert_col(TotalArray[i],ArrayCol)
        TotalArray[i]=array_set(TotalArray[i], 0, ArrayCol, "负责人")   #插入
        TotalArray[i]=insert_col(TotalArray[i],ArrayCol)
        TotalArray[i]=array_set(TotalArray[i], 0, ArrayCol, "供应商")   #插入
        InvPosition = ArrayCol+2                                #可用库存位置
        OomPosition = ArrayCol+3                                #缺料时间位置
        PlanPosition = array_col(TotalArray[i])                 #计划开始位置
        set_Prog(self,row_i,MaxRow)
        win.update()#更新界面
        #导入计划表
        print_label(self,"插入处理计划表")
        print_log(self,"插入处理计划表")
        TotalArray[i]=insert_row(TotalArray[i],0)#插入行
        TotalArray[i]=insert_row(TotalArray[i],0)#插入行
        TotalArray[i]=array_set(TotalArray[i], 0, (PlanPosition-7), PlanSheet["F4"].value)#迈腾成品料号
        TotalArray[i]=array_set(TotalArray[i], 0, (PlanPosition-6), PlanSheet.cell(4,TypeList[i]).value)#料号
        TotalArray[i]=array_set(TotalArray[i], 0, (PlanPosition-5), PlanSheet["B4"].value)#品名
        TotalArray[i]=array_set(TotalArray[i], 0, (PlanPosition-4), PlanSheet["H4"].value)#FCS
        TotalArray[i]=array_set(TotalArray[i], 0, (PlanPosition-3), PlanSheet["I4"].value)#已完成数
        TotalArray[i]=array_set(TotalArray[i], 0, (PlanPosition-2), PlanSheet["J4"].value)#未完成数
        TotalArray[i]=array_set(TotalArray[i], 0, (PlanPosition-1), PlanSheet["K4"].value)#计划数
        PlanModelPosition = PlanPosition-6
        for Col_i in range(PlanSheet.max_column-11):#12
            if((PlanSheet.cell(4,Col_i+12).value!=None)and(PlanSheet.cell(5,Col_i+12).value!=None)):
                TotalArray[i]=array_set(TotalArray[i], 0, Col_i+PlanPosition, excel_to_date(PlanSheet.cell(4,Col_i+12).value)[0])
                TotalArray[i]=array_set(TotalArray[i], 1, Col_i+PlanPosition, excel_to_date(PlanSheet.cell(5,Col_i+12).value)[2])
                TotalArray[i]=array_set(TotalArray[i], 2, Col_i+PlanPosition, excel_to_date(PlanSheet.cell(4,Col_i+12).value)[1])
            else:
                break
        for Row_i in range(PlanSheet.max_row-6):
            set_Prog(self,Row_i,PlanSheet.max_row-6)
            win.update()#更新界面
            if(PlanSheet.cell(PlanSheet.max_row-Row_i,TypeList[i]).value!=None):
                TotalArray[i]=insert_row(TotalArray[i],2)#插入行
                TotalArray[i]=array_set(TotalArray[i], 2, (PlanPosition-7), PlanSheet.cell(PlanSheet.max_row-Row_i,6).value)#迈腾成品料号
                TotalArray[i]=array_set(TotalArray[i], 2, (PlanPosition-6), PlanSheet.cell(PlanSheet.max_row-Row_i,TypeList[i]).value)#料号
                TotalArray[i]=array_set(TotalArray[i], 2, (PlanPosition-5), PlanSheet.cell(PlanSheet.max_row-Row_i,2).value)#品名
                TotalArray[i]=array_set(TotalArray[i], 2, (PlanPosition-4), PlanSheet.cell(PlanSheet.max_row-Row_i,8).value)#FCS
                TotalArray[i]=array_set(TotalArray[i], 2, (PlanPosition-3), PlanSheet.cell(PlanSheet.max_row-Row_i,9).value)#已完成数
                TotalArray[i]=array_set(TotalArray[i], 2, (PlanPosition-2), PlanSheet.cell(PlanSheet.max_row-Row_i,10).value)#未完成数
                TotalArray[i]=array_set(TotalArray[i], 2, (PlanPosition-1), PlanSheet.cell(PlanSheet.max_row-Row_i,11).value)#计划数
                for Col_i in range(PlanSheet.max_column-13):#12
                    if((PlanSheet.cell(4,Col_i+12).value!=None)or(PlanSheet.cell(5,Col_i+12).value!=None)):
                        TotalArray[i]=array_set(TotalArray[i], 2, Col_i+PlanPosition, PlanSheet.cell(PlanSheet.max_row-Row_i,Col_i+12).value)#计划
                    else:
                        break
                PlanMaxRow[i]+=1
        PlanWorkbook.close()
        #处理日用量
        for Row_i in range(PlanMaxRow[i]+2,array_row(TotalArray[i])):
            TotalArray[i] = array_set(TotalArray[i],Row_i,(PlanPosition-1),"需求")
            for Col_i in range(PlanPosition,array_col(TotalArray[i])):
                TotalArray[i] = array_set(TotalArray[i],Row_i,Col_i,0)
        for Plan_i in range(2,PlanMaxRow[i]):#计划遍历
            PlanModelName=array_get(TotalArray[i],Plan_i,PlanModelPosition)
            print_label(self,"处理计划:"+str(Plan_i-1)+"/"+str(PlanMaxRow[i]-2))
            print_log(self,"处理计划:"+PlanModelName)
            set_Prog(self,Plan_i,PlanMaxRow[i])
            win.update()#更新界面
            for Modell_i in range(UsePosition):#bom遍历
                ModelName=array_get(TotalArray[i],PlanMaxRow[i],Modell_i)
                if(PlanModelName==ModelName):
                    for Col_i in range(PlanPosition,array_col(TotalArray[i])):
                        PlannedQuantity = array_get(TotalArray[i],Plan_i,Col_i)#型号当天计划量
                        if(PlannedQuantity!=None):
                            for Row_i in range(PlanMaxRow[i]+2,array_row(TotalArray[i])):
                                SingleDosage = array_get(TotalArray[i],Row_i,UsePosition+1+Modell_i)#物料用量
                                if(SingleDosage==None):
                                    SingleDosage = 0
                                UseRatio = array_get(TotalArray[i],Row_i,UsePosition)#使用比例
                                TotalDosage = array_get(TotalArray[i],Row_i,Col_i)#物料当天用量
                                TotalDosage +=UseRatio*SingleDosage*PlannedQuantity
                                TotalArray[i] = array_set(TotalArray[i],Row_i,Col_i,TotalDosage)
                    break
            else:
                print_log(self," └错误:BOM中无法找到此料号!")
        #插入行
        ArrayRow = array_row(TotalArray[i])#获取最大行数
        ArrayCol = array_col(TotalArray[i])#获取最大列数
        ArrayRowNum = ArrayRow-PlanMaxRow[i]-2
        print_log(self,"插入行...")
        for Row_i in range(ArrayRowNum):
            print_label(self,"处理计划:"+str(Row_i+1)+"/"+str(ArrayRowNum))
            set_Prog(self,Row_i,ArrayRowNum)
            win.update()#更新界面
            TotalArray[i] = insert_row(TotalArray[i], ArrayRow-Row_i)
            TotalArray[i] = array_set(TotalArray[i], ArrayRow-Row_i, PlanPosition-1,"GAP")
            TotalArray[i] = insert_row(TotalArray[i], ArrayRow-Row_i)
            TotalArray[i] = array_set(TotalArray[i], ArrayRow-Row_i, PlanPosition-1,"ETA")
        #增加GAP公式，增加缺料时间公式
        print_log(self,"增加公式...")
        for Row_i in range(ArrayRowNum):
            print_label(self,"增加公式:"+str(Row_i+1)+"/"+str(ArrayRowNum))
            set_Prog(self,Row_i,ArrayRowNum)
            win.update()#更新界面
            ReqRow = PlanMaxRow[i]+2+Row_i*3
            GapOldXc = CxTransform(ReqRow+1,InvPosition+1)#剩余库存坐标
            EtaXc = CxTransform(ReqRow+2,PlanPosition+1)#ETA坐标
            DemandXc = CxTransform(ReqRow+1,PlanPosition+1)#需求坐标
            formulastr='='+GapOldXc+'+'+EtaXc+'-'+DemandXc
            TotalArray[i] = array_set(TotalArray[i], ReqRow+2, PlanPosition,formulastr)#填入公式
            for Col_i in range((PlanPosition+1),(ArrayCol)):
                GapOldXc = CxTransform(ReqRow+3,Col_i)#剩余库存坐标
                EtaXc = CxTransform(ReqRow+2,Col_i+1)#ETA坐标
                DemandXc = CxTransform(ReqRow+1,Col_i+1)#需求坐标
                formulastr='='+GapOldXc+'+'+EtaXc+'-'+DemandXc
                TotalArray[i] = array_set(TotalArray[i], ReqRow+2, Col_i,formulastr)#填入公式

            #生成缺料时间公式
            OomStr = "=IF("
            for Col_i in range((PlanPosition),(ArrayCol-1)):
                GapXc = CxTransform(ReqRow+3,Col_i+1)#GAP坐标
                dateXc = CxTransform(PlanMaxRow[i]+1,Col_i+1)#日期坐标
                OomStr+=GapXc+"<0,"+dateXc+",IF("
            else :
                GapXc = CxTransform(ReqRow+3,ArrayCol)#GAP坐标
                dateXc = CxTransform(PlanMaxRow[i]+1,ArrayCol)#日期坐标
                OomStr+=GapXc+"<0,"+dateXc+",\"No\""
            for Col_i in range((PlanPosition),(ArrayCol)):
                OomStr+=")"
            TotalArray[i] = array_set(TotalArray[i], ReqRow, OomPosition,OomStr)
    #存储表格
    TotalArrayLen = [array_row(TotalArray[0]),array_row(TotalArray[1]),array_row(TotalArray[2])]
    ProgLen = TotalArrayLen[0]+TotalArrayLen[1]+TotalArrayLen[2]
    set_Prog(self,0,ProgLen)#更新进度条
    print_label(self,"处理MT料况表")
    print_log(self,"处理MT料况表")
    win.update()#更新界面
    # 创建一个新的Excel文件
    OutWorkBook = openpyxl.Workbook()
    # 选择一个工作表
    OutSheet = (OutWorkBook.create_sheet("电子料",0),
               OutWorkBook.create_sheet("结构料",1),
               OutWorkBook.create_sheet("包材",2))
    for i in range(3):
        for j in range(TotalArrayLen[i]):
            if i == 0:
                ProgBar = j
            elif i == 1:
                ProgBar = TotalArrayLen[0]+j
            else :
                ProgBar = TotalArrayLen[0]+TotalArrayLen[0]+j
            set_Prog(self,ProgBar,ProgLen)#更新进度条
            win.update()#更新界面
            OutSheet[i].append(TotalArray[i][j])

    #TotalArray.close()#释放内存
    
    '''
    for i in range(3):
        
        print_log(self,"生成格式:"+str(i+1)+"/"+str(3))
        win.update()#更新界面
        RowStart = 0
        ColStart = 0
        RowMax = OutSheet[i].max_row
        ColMax = OutSheet[i].max_column
        align = Alignment(horizontal='center', vertical='center' , wrapText=True)
        for Row_i in range(1,OutSheet[i].max_row):#索引开始行（BOM）
            if(OutSheet[i].cell(Row_i,1).value!=None):
                RowStart = Row_i
                break
        for Col_i in range(1,OutSheet[i].max_column):#索引开始列（计划）
            if(OutSheet[i].cell(1,Col_i).value!=None):
                ColStart = Col_i
                break
        for Col_i in range(1,ColStart-2):
            OutSheet[i].column_dimensions[get_column_letter(Col_i)].width = 4
        for Col_i in range(ColStart-3,ColMax+1):
            StartCx = CxTransform(RowStart,Col_i)
            endCx = CxTransform(RowStart+1,Col_i)
            OutSheet[i].merge_cells(StartCx+':'+endCx)  # 合并单元格
            OutSheet[i][StartCx].alignment = align      # 设置水平和垂直对齐方式为居中
        MaterialMax = int((RowMax - RowStart - 1)/3)
        SerialNumOld = 0
        StartCx = CxTransform(1,1)
        for Row_i in range(MaterialMax):
            SerialNumNew = OutSheet[i].cell(RowStart+2+Row_i*3,ColStart-3).value
            if(SerialNumOld!=SerialNumNew):
                if(SerialNumOld!=0):
                    endCx = CxTransform(RowStart+1+Row_i*3,ColStart-3)
                    OutSheet[i].merge_cells(StartCx+':'+endCx)  # 合并单元格
                    OutSheet[i][StartCx].alignment = align      # 设置水平和垂直对齐方式为居中
                StartCx = CxTransform(RowStart+2+Row_i*3,ColStart-3)
                SerialNumOld=SerialNumNew
        else:
            endCx = CxTransform(RowStart+1+MaterialMax*3,ColStart-3)
            OutSheet[i].merge_cells(StartCx+':'+endCx)  # 合并单元格
            OutSheet[i][StartCx].alignment = align      # 设置水平和垂直对齐方式为居中
        #####合并表格
        for Col_i in range(ColStart-2,ColStart+6):
            print_label(self,"格式:"+str(Col_i+1)+"/"+str(ColStart+6))
            for Row_i in range(MaterialMax):
                set_Prog(self,Row_i+1,MaterialMax)#更新进度条
                win.update()#更新界面
                StartCx = CxTransform(RowStart+2+Row_i*3,Col_i)
                endCx = CxTransform(RowStart+4+Row_i*3,Col_i)
                OutSheet[i].merge_cells(StartCx+':'+endCx)  # 合并单元格
                OutSheet[i][StartCx].alignment = align      # 设置水平和垂直对齐方式为居中
    '''
    
    '''
    if(self.checkbox_var.get()==1):  
        for i in range(3):
            for Col_i in range(1,ColStart+5):
                set_Prog(self,Col_i,ColStart+5)#更新进度条
                win.update()#更新界面
                for Row_i in range(MaterialMax):
                    StartCx = CxTransform(RowStart+2+Row_i*3,Col_i)
                    endCx = CxTransform(RowStart+4+Row_i*3,Col_i)
                    OutSheet[i].merge_cells(StartCx+':'+endCx)  # 合并单元格
                    OutSheet[i][StartCx].alignment = align      # 设置水平和垂直对齐方式为居中
    '''
            
            

        
        
            
            
    print_label(self,"存储MT料况表")
    print_log(self,"存储MT料况表")
    win.update()#更新界面
    try:
        # 获取当前日期
        current_date = str(datetime.date.today().strftime("%Y%m%d"))
        SaveName=".\MT料况表"+current_date+".xlsx"
        OutWorkBook.save(SaveName)
    except:
        current_datetime = str(datetime.datetime.now().strftime("%Y%m%d%H%M%S"))
        SaveName=".\MT料况表"+current_datetime+".xlsx"
        OutWorkBook.save(SaveName)
        
    OutWorkBook.close()
    print_label(self,"保存："+SaveName)
    print_log(self,"保存："+SaveName)
    win.update()#更新界面

#UI#######################################################################

class WinGUI(Tk):
    def __init__(self):
        super().__init__()
        self.__win()
        self.checkbox_var = IntVar(value=0)
        self.tk_label_lm3lm3mh = self.__tk_label_lm3lm3mh(self)
        self.tk_input_lm3ln8cp = self.__tk_input_lm3ln8cp(self)
        self.tk_button_lm3lnzud = self.__tk_button_lm3lnzud(self)
        self.tk_label_lm3lohwd = self.__tk_label_lm3lohwd(self)
        self.tk_input_lm3lqahh = self.__tk_input_lm3lqahh(self)
        self.tk_button_lm3lqeut = self.__tk_button_lm3lqeut(self)
        self.tk_label_lm3lw72f = self.__tk_label_lm3lw72f(self)
        self.tk_check_button_lm3lwui9 = self.__tk_check_button_lm3lwui9(self)
        self.tk_button_lm3lx0ji = self.__tk_button_lm3lx0ji(self)
        self.tk_progressbar_lm3lze9x = self.__tk_progressbar_lm3lze9x(self)
        self.tk_text_lm3m3ylm = self.__tk_text_lm3m3ylm(self)
    def __win(self):
        self.title("料况输出V2.1极速版")
        # 设置窗口大小、居中
        width = 600
        height = 290
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        geometry = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(geometry)
        self.resizable(width=False, height=False)
        # 自动隐藏滚动条
    def scrollbar_autohide(self,bar,widget):
        self.__scrollbar_hide(bar,widget)
        widget.bind("<Enter>", lambda e: self.__scrollbar_show(bar,widget))
        bar.bind("<Enter>", lambda e: self.__scrollbar_show(bar,widget))
        widget.bind("<Leave>", lambda e: self.__scrollbar_hide(bar,widget))
        bar.bind("<Leave>", lambda e: self.__scrollbar_hide(bar,widget))
    
    def __scrollbar_show(self,bar,widget):
        bar.lift(widget)
    def __scrollbar_hide(self,bar,widget):
        bar.lower(widget)
    
    def vbar(self,ele, x, y, w, h, parent):
        sw = 15 # Scrollbar 宽度
        x = x + w - sw
        vbar = Scrollbar(parent)
        ele.configure(yscrollcommand=vbar.set)
        vbar.config(command=ele.yview)
        vbar.place(x=x, y=y, width=sw, height=h)
        self.scrollbar_autohide(vbar,ele)
    def __tk_label_lm3lm3mh(self,parent):
        label = Label(parent,text="BOM清单：",anchor="center", )
        label.place(x=20, y=20, width=80, height=30)
        return label
    def __tk_input_lm3ln8cp(self,parent):
        ipt = Entry(parent, )
        ipt.place(x=100, y=20, width=420, height=30)
        return ipt
    def __tk_button_lm3lnzud(self,parent):
        btn = Button(parent, text="选择", takefocus=False,)
        btn.place(x=530, y=20, width=50, height=30)
        return btn
    def __tk_label_lm3lohwd(self,parent):
        label = Label(parent,text="生产计划表：",anchor="center", )
        label.place(x=20, y=60, width=80, height=30)
        return label
    def __tk_input_lm3lqahh(self,parent):
        ipt = Entry(parent, )
        ipt.place(x=100, y=60, width=420, height=30)
        return ipt
    def __tk_button_lm3lqeut(self,parent):
        btn = Button(parent, text="选择", takefocus=False,)
        btn.place(x=530, y=60, width=50, height=30)
        return btn
    def __tk_label_lm3lw72f(self,parent):
        label = Label(parent,text="请选择文件并开始",anchor="w", )
        label.place(x=20, y=100, width=360, height=30)
        return label
    def __tk_check_button_lm3lwui9(self,parent):
        cb = Checkbutton(parent,text="生成格式",variable=self.checkbox_var,state=DISABLED)
        cb.place(x=400, y=100, width=80, height=30)
        return cb
    def __tk_button_lm3lx0ji(self,parent):
        btn = Button(parent, text="开始生成", takefocus=False,)
        btn.place(x=500, y=100, width=80, height=30)
        return btn
    def __tk_progressbar_lm3lze9x(self,parent):
        progressbar = Progressbar(parent, orient=HORIZONTAL,)
        progressbar.place(x=20, y=140, width=560, height=10)
        return progressbar
    def __tk_text_lm3m3ylm(self,parent):
        text = Text(parent,state="disabled")
        text.place(x=20, y=170, width=560, height=100)
        self.vbar(text, 20, 170, 560, 100,parent)
        return text
class Win(WinGUI):
    def __init__(self):
        super().__init__()
        self.__event_bind()
    def OpenBomEvent(self):#(self,evt):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if(file_path!=""):
            self.tk_input_lm3ln8cp.delete(0, END)  # 清空输入框内容
            self.tk_input_lm3ln8cp.insert(END, file_path)  # 将选择的目录路径填入输入框
    def OpenPlanEvent(self):#(self,evt):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if(file_path!=""):
            self.tk_input_lm3lqahh.delete(0, END)  # 清空输入框内容
            self.tk_input_lm3lqahh.insert(END, file_path)  # 将选择的目录路径填入输入框
    def StartProcessEvent(self):#(self,evt):
        self.tk_button_lm3lx0ji.config(state=DISABLED)
        #self.tk_check_button_lm3lwui9.config(state=DISABLED)
        win.update()#更新界面
        main_app(self)
        #self.tk_check_button_lm3lwui9.config(state=NORMAL)
        self.tk_button_lm3lx0ji.config(state=NORMAL)
        
    def __event_bind(self):
        #self.tk_button_lm3lnzud.bind('<Button>',self.OpenBomEvent)
        #self.tk_button_lm3lqeut.bind('<Button>',self.OpenPlanEvent)
        #self.tk_button_lm3lx0ji.bind('<Button>',self.StartProcessEvent)
        self.tk_button_lm3lnzud["command"]=self.OpenBomEvent
        self.tk_button_lm3lqeut["command"]=self.OpenPlanEvent
        self.tk_button_lm3lx0ji["command"]=self.StartProcessEvent
        pass
if __name__ == "__main__":
    win = Win()
    win.mainloop()