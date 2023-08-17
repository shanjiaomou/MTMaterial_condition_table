import openpyxl 

#设置单元格格式
cell_format = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
#设置单元格边
thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'), 
                                             right=openpyxl.styles.borders.Side(style='thin'), 
                                             top=openpyxl.styles.borders.Side(style='thin'), 
                                             bottom=openpyxl.styles.borders.Side(style='thin'))

# 打开Excel文件
BomWorkbook = openpyxl.load_workbook('BOM使用比例清单.xlsx')
try:
    NeedWorkbook = openpyxl.load_workbook('小米生产主计划.xlsx')

except:
    print("请准备文件：小米生产主计划.xlsx")
    exit()
NewWorkbook = openpyxl.Workbook()   #新表格

SpaceSize=5

# 选择一个工作表
BomSheet = (BomWorkbook["电子料"],
            BomWorkbook["结构料"],
            BomWorkbook["包材"])
NewSheet = (NewWorkbook.create_sheet("电子料",0),
            NewWorkbook.create_sheet("结构料",1),
            NewWorkbook.create_sheet("包材",2))
NeedSheet = NeedWorkbook["小米生产计划"]
#NeedSheet.calculate()
if(NeedSheet.cell(4, 3).value!="整机料号")\
or(NeedSheet.cell(4, 4).value!="半成品料号")\
or(NeedSheet.cell(4, 5).value!="PCBA"):
    print("计划表格式不对！")
    exit()

PlannedHeight=NeedSheet.max_row-4

offset = 0
for i in range(1,BomSheet[0].max_column):
    if(BomSheet[0].cell(1, i).value=="替代料"):
        offset = i-1

for i in range(3):
    NewSheet[i].cell(1,offset*2+5+SpaceSize).value = NeedSheet.cell(4, 6).value     #迈腾成品料号 4 6

    NewSheet[i].cell(1,offset*2+5+SpaceSize+1).value = NeedSheet.cell(4, 5-i).value #整机/半成品/PCBA

    NewSheet[i].cell(1,offset*2+5+SpaceSize+2).value = NeedSheet.cell(4, 2).value   #机型

    for j in range(8,12):
        NewSheet[i].cell(1,offset*2+SpaceSize+j).value = NeedSheet.cell(4, j).value
    
    for j in range(offset*2+5+SpaceSize,offset*2+12+SpaceSize):
        NewSheet[i].cell(1, j).border = thin_border #边框
        NewSheet[i].column_dimensions[openpyxl.utils.get_column_letter(j)].width = 10 #列宽
        NewSheet[i].cell(1, j).alignment = cell_format #居中
        NewSheet[i].merge_cells(openpyxl.utils.get_column_letter(j)+"1:"+openpyxl.utils.get_column_letter(j)+"2")


    for j in range(12,NeedSheet.max_column):
        NewSheet[i].column_dimensions[openpyxl.utils.get_column_letter(offset*2+SpaceSize+j)].width = 10 #列宽
        NewSheet[i].cell(1,offset*2+SpaceSize+j).value = NeedSheet.cell(4, j).value
        NewSheet[i].cell(1, offset*2+SpaceSize+j).border = thin_border #边框
        NewSheet[i].cell(1, offset*2+SpaceSize+j).alignment = cell_format #居中
        NewSheet[i].cell(1, offset*2+SpaceSize+j).number_format = 'm月d日'
        NewSheet[i].cell(2,offset*2+SpaceSize+j).value = NeedSheet.cell(5, j).value
        NewSheet[i].cell(2, offset*2+SpaceSize+j).border = thin_border #边框
        NewSheet[i].cell(2, offset*2+SpaceSize+j).alignment = cell_format #居中
        NewSheet[i].cell(2, offset*2+SpaceSize+j).number_format = '(aaa)'
    NewSheet[i].row_dimensions[1].height = 12.8 #设置行高
    NewSheet[i].row_dimensions[2].height = 12.8 #设置行高

    for k in range(6,NeedSheet.max_row):
        NewSheet[i].cell(k-3,offset*2+5+SpaceSize).value = NeedSheet.cell(k, 6).value     #迈腾成品料号 4 6
        NewSheet[i].cell(k-3,offset*2+5+SpaceSize+1).value = NeedSheet.cell(k, 5-i).value #整机/半成品/PCBA
        NewSheet[i].cell(k-3,offset*2+5+SpaceSize+2).value = NeedSheet.cell(k, 2).value   #机型
        for j in range(3):
            NewSheet[i].cell(k-3, offset*2+5+SpaceSize+j).border = thin_border #边框
            NewSheet[i].cell(k-3, offset*2+5+SpaceSize+j).alignment = cell_format #居中
        for j in range(8,NeedSheet.max_column):
            NewSheet[i].cell(k-3,offset*2+SpaceSize+j).value = NeedSheet.cell(k, j).value
            NewSheet[i].cell(k-3, offset*2+SpaceSize+j).border = thin_border #边框
            NewSheet[i].cell(k-3, offset*2+SpaceSize+j).alignment = cell_format #居中
        NewSheet[i].row_dimensions[k-3].height = 12.8 #设置行高
    #复制bom表
    for j in range(1,BomSheet[i].max_column+1):
        if((j<=(offset+1))or((j>=(offset+5))and(j<=(offset*2+6)))):
            NewSheet[i].column_dimensions[openpyxl.utils.get_column_letter(j)].width = 4
        elif(j==(offset+2)):
            NewSheet[i].column_dimensions[openpyxl.utils.get_column_letter(j)].width = 15
        elif(j==(offset+3)):
            NewSheet[i].column_dimensions[openpyxl.utils.get_column_letter(j)].width = 18
        elif(j==(offset+4)):
            NewSheet[i].column_dimensions[openpyxl.utils.get_column_letter(j)].width = 9
        for k in range(1,BomSheet[i].max_row+1):
            NewSheet[i].cell(k+PlannedHeight,j).value = BomSheet[i].cell(k,j).value
            NewSheet[i].cell(k+PlannedHeight,j).border = thin_border #边框
            NewSheet[i].cell(k+PlannedHeight,j).alignment = cell_format #居中
        if(((j>=(offset+1))and(j<=(offset+5)))or(j==(offset*2+6))):#合并单元格
            setbuf=(openpyxl.utils.get_column_letter(j)+str(PlannedHeight+1)+":"+openpyxl.utils.get_column_letter(j)+str(PlannedHeight+2))
            NewSheet[i].merge_cells(setbuf)

    NewSheet[i].cell(1+PlannedHeight,offset*2+10+SpaceSize).value = "共用料"
    NewSheet[i].cell(1+PlannedHeight,offset*2+11+SpaceSize).value = "区分"



# 保存文件
NewWorkbook.save(".\MT料况表.xlsx")
NewWorkbook.close()
